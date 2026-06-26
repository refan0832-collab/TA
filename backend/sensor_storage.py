import sqlite3
import os
import threading
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =========================
# PATH
# =========================

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DB_FILE     = os.path.join(BASE_DIR, "sensor_data.db")
EXPORT_DIR  = os.path.join(BASE_DIR, "..", "exports")

_lock = threading.Lock()

# =========================
# INIT DB
# =========================

def init_db():

    os.makedirs(EXPORT_DIR, exist_ok=True)

    with sqlite3.connect(DB_FILE) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sensor_history (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp   TEXT NOT NULL,
                tegangan    REAL,
                arus        REAL,
                daya        REAL,
                frekuensi   REAL,
                pf          REAL,
                energy      REAL,
                overvoltage INTEGER DEFAULT 0,
                undervoltage INTEGER DEFAULT 0
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_timestamp
            ON sensor_history(timestamp)
        """)
        conn.commit()

    print("✅ SQLite sensor_data.db siap")

# =========================
# SIMPAN DATA
# =========================

def save_sensor(data):

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with _lock:
        with sqlite3.connect(DB_FILE) as conn:
            conn.execute("""
                INSERT INTO sensor_history
                (timestamp, tegangan, arus, daya, frekuensi, pf, energy, overvoltage, undervoltage)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                timestamp,
                data.get("tegangan",     0),
                data.get("arus",         0),
                data.get("daya",         0),
                data.get("frekuensi",    0),
                data.get("pf",           0),
                data.get("energy",       0),
                1 if data.get("overvoltage",  False) else 0,
                1 if data.get("undervoltage", False) else 0,
            ))
            conn.commit()

# =========================
# AMBIL DATA HARI INI
# =========================

def get_today():

    today = datetime.now().strftime("%Y-%m-%d")

    with _lock:
        with sqlite3.connect(DB_FILE) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute("""
                SELECT timestamp, tegangan, arus, daya, frekuensi, pf
                FROM sensor_history
                WHERE timestamp >= ?
                ORDER BY timestamp ASC
            """, (f"{today} 00:00:00",)).fetchall()

    return [dict(r) for r in rows]

# =========================
# AMBIL DATA RANGE TANGGAL
# =========================

def get_by_date(date_str):

    with _lock:
        with sqlite3.connect(DB_FILE) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute("""
                SELECT timestamp, tegangan, arus, daya, frekuensi, pf, energy
                FROM sensor_history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY timestamp ASC
            """, (
                f"{date_str} 00:00:00",
                f"{date_str} 23:59:59"
            )).fetchall()

    return [dict(r) for r in rows]

# =========================
# RESET DATA BY DATE
# =========================

def reset_by_date(date_str):

    with _lock:
        with sqlite3.connect(DB_FILE) as conn:
            deleted = conn.execute("""
                DELETE FROM sensor_history
                WHERE timestamp >= ? AND timestamp < ?
            """, (
                f"{date_str} 00:00:00",
                f"{date_str} 23:59:59"
            )).rowcount
            conn.commit()

    print(f"🗑️  Reset data sensor {date_str}: {deleted} record dihapus")
    return deleted

# =========================
# DAFTAR TANGGAL TERSEDIA
# =========================

def get_available_dates():

    with _lock:
        with sqlite3.connect(DB_FILE) as conn:
            rows = conn.execute("""
                SELECT DISTINCT DATE(timestamp) as date
                FROM sensor_history
                ORDER BY date DESC
            """).fetchall()

    return [r[0] for r in rows]

# =========================
# AUTO CLEANUP — hapus data > 7 hari
# Dipanggil setiap hari sebelum export
# =========================

def cleanup_old_data():

    cutoff = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")

    with _lock:
        with sqlite3.connect(DB_FILE) as conn:
            deleted = conn.execute("""
                DELETE FROM sensor_history
                WHERE timestamp < ?
            """, (cutoff,)).rowcount
            conn.commit()

    if deleted > 0:
        print(f"🗑️  Cleanup: {deleted} record lama dihapus (> 7 hari)")

# =========================
# EXPORT EXCEL HARIAN
# Dipanggil otomatis tiap ganti hari
# =========================

def export_daily_excel(date_str=None):

    if date_str is None:
        # Export kemarin
        date_str = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

    export_path = os.path.join(
        EXPORT_DIR,
        f"sensor_{date_str}.xlsx"
    )

    # Kalau sudah ada, skip
    if os.path.exists(export_path):
        print(f"⏭️  Export {date_str} sudah ada, skip")
        return export_path

    rows = get_by_date(date_str)

    if not rows:
        print(f"⚠️  Tidak ada data untuk {date_str}, skip export")
        return None

    # =========================
    # BUAT EXCEL
    # =========================

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sensor {date_str}"

    # HEADER STYLE
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "Timestamp",
        "Tegangan (V)",
        "Arus (A)",
        "Daya (W)",
        "Frekuensi (Hz)",
        "Power Factor"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align

    # LEBAR KOLOM
    col_widths = [22, 14, 10, 10, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ISI DATA
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r["timestamp"])
        ws.cell(row=row_idx, column=2, value=round(r["tegangan"],  2))
        ws.cell(row=row_idx, column=3, value=round(r["arus"],      2))
        ws.cell(row=row_idx, column=4, value=round(r["daya"],      2))
        ws.cell(row=row_idx, column=5, value=round(r["frekuensi"], 2))
        ws.cell(row=row_idx, column=6, value=round(r["pf"],        2))

        # Warna baris selang-seling
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="F0F4F8")
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = fill

    wb.save(export_path)
    print(f"✅ Export Excel: sensor_{date_str}.xlsx ({len(rows)} record)")
    return export_path

# =========================
# DAFTAR FILE EXCEL TERSEDIA
# =========================

def get_export_files():

    os.makedirs(EXPORT_DIR, exist_ok=True)

    files = []
    for f in sorted(os.listdir(EXPORT_DIR), reverse=True):
        if f.startswith("sensor_") and f.endswith(".xlsx"):
            date_str = f.replace("sensor_", "").replace(".xlsx", "")
            size_kb  = os.path.getsize(
                os.path.join(EXPORT_DIR, f)
            ) // 1024
            files.append({
                "filename": f,
                "date":     date_str,
                "size_kb":  size_kb
            })

    return files